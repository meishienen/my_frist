# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   jiekouzidonghua
# FileName:     read_excell
# Author:      liuyeyu
# Datetime:    2020/10/14 17:38
# Description:
#------------------------------------------------------------------------------------
import openpyxl

from InterfaceTest.common.read_json import read_json
from InterfaceTest.common.read_ini import ReadIni
from InterfaceTest.data.excel_column import ExcelColumn

class ReadExcel:
    def __init__(self):
        self.readini=ReadIni()
        self.json_params_path=self.readini.params_path()
        self.json_except_path=self.readini.except_path()
        #打开表格
        self.wb=openpyxl.load_workbook(self.readini.excel_path())
        #指定Sheet页
        self.ws=self.wb[self.readini.Sheet_path()]

    #指定方法获取单元格的值
    def get_cell_value(self,column,row):
        return self.ws[column+str(row)].value

    #获取用例编号
    def get_case_id(self,row):
        return self.get_cell_value(ExcelColumn.CASE_ID,row)

    #获取用例标题
    def get_case_title(self,row):
        return self.get_cell_value(ExcelColumn.CASE_TITLE,row)

    #获取用例请求
    def get_case_method(self,row):
        return self.get_cell_value(ExcelColumn.CASE_METHOD,row)

    #获取用例URL
    def get_case_url(self, row):
        return self.get_cell_value(ExcelColumn.CASE_URL, row)

    # 获取用例params(json文件中的key值)
    def get_case_params_key(self, row):
        return self.get_cell_value(ExcelColumn.CASE_PARAMS, row)
    #获取用例params(json文件中的value值)
    def get_case_params_value(self, row):
        key=self.get_case_params_key(row)
        if key:
            return read_json(self.json_params_path)[key]

    #获取用例期望结果(json文件中的key值)
    def get_case_except_key(self, row):
        return self.get_cell_value(ExcelColumn.CASE_EXCEPT, row)
    # 获取用例期望结果(json文件中的value值)
    def get_case_except_value(self, row):
        k_key=self.get_case_except_key(row)
        if k_key:
            return read_json(self.json_except_path)[k_key]

    #获取用例行数
    def get_row_count(self):
        return self.ws.max_row

    # 获取前置用例编号
    def get_case_precondition_id(self, row):
        return self.get_cell_value(ExcelColumn.CASE_PRECONDITION_ID, row)

    # 获取正则表达式
    def get_re(self, row):
        return self.get_cell_value(ExcelColumn.CASE_RE, row)

    # 获取依赖字段
    def get_case_depend_key(self, row):
        return self.get_cell_value(ExcelColumn.CASE_DEPEND_KEY, row)

    # 获取是否执行的字段
    def get_case_if_execute(self, row):
        return self.get_cell_value(ExcelColumn.CASE_IF_EXECUTE, row)

    #获取响应格式
    def get_case_type(self, row):
        return self.get_cell_value(ExcelColumn.CASE_TYPE, row)

if __name__ == '__main__':
    excel = ReadExcel()
    row = 2
    print(excel.get_cell_value("A",2))


